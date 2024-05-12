from openpyxl import load_workbook
from modelos.operacion import Operacion  
from modelos.broker import BrokerEnum
from decimal import Decimal
from modelos.tipo_operacion import TipoOperacion
from datetime import date, datetime
from servicios.id_generator import gen_id
def leer_excel_evo_y_mapear_objetos(ruta_archivo,broker):
    libro = load_workbook(filename=ruta_archivo, data_only=True)
    hoja = libro.active
    operaciones = []
    rows_values = list(hoja.iter_rows(min_row=2, values_only=True))[1:]

    for fila in rows_values:
        fecha_operacion_str,fecha_liquidacion,id_operacion_broker,mercado,operacion_str,isin,valor,títulos_nominal,divisa,precio_neto,importe_neto = fila
        operacion = _parse_operacion(operacion_str)
        fecha_operacion = parse_evo_date(fecha_operacion_str)
        operacion = Operacion(
            id=gen_id(),
            fecha=fecha_operacion, 
            isin=isin, 
            tipo=operacion, 
            cantidad=títulos_nominal, 
            precio_unitario=Decimal(precio_neto), 
            divisa=divisa, 
            nombre=valor, 
            importe_neto=Decimal(importe_neto), 
            broker=broker
        )
        operaciones.append(operacion)
    
    return operaciones

def _parse_operacion(operacion_str: str) -> TipoOperacion:
    if not operacion_str:
        return None
    stripped_op = operacion_str.strip().lower()
    if stripped_op == 'compra':
        return TipoOperacion.COMPRA
    if stripped_op == 'venta':
        return TipoOperacion.VENTA
    if stripped_op == 'dividendo':
        return TipoOperacion.DIVIDENDO

def parse_evo_date(str_date):
    return datetime.strptime( str_date,"%Y-%m-%d").date()
